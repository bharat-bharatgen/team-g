"""
Shared Excel generation utilities.

Provides common styles, constants, and generic helpers for building
.xlsx workbooks across different services (MER, pathology, risk, test verification).
"""

import io
from typing import Any, Callable, Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet


# ─── Common Styles ───────────────────────────────────────────────────────────

FILL_WHITE = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_YELLOW = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
FILL_RED = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
FILL_GRAY = PatternFill(start_color="E5E5E5", end_color="E5E5E5", fill_type="solid")
FILL_LIGHT_BLUE = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
FILL_ORANGE = PatternFill(start_color="FFD6A5", end_color="FFD6A5", fill_type="solid")
FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

FONT_HEADER = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
FONT_DEFAULT = Font(name="Calibri", size=11)
FONT_BOLD = Font(name="Calibri", size=11, bold=True)
FONT_RED_BOLD = Font(name="Calibri", size=11, bold=True, color="DC2626")
FONT_TITLE = Font(name="Calibri", size=14, bold=True)
FONT_HIDDEN = Font(size=1)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


# ─── Column specification ────────────────────────────────────────────────────

class ColSpec:
    """Describes a single visible column in a sheet."""

    __slots__ = ("header", "width", "align")

    def __init__(self, header: str, width: int = 15, align: str = "left"):
        self.header = header
        self.width = width
        self.align = align


# ─── Row style result ────────────────────────────────────────────────────────

class RowStyle:
    """Fill / font overrides for a single data row."""

    __slots__ = ("fill", "font_overrides")

    def __init__(
        self,
        fill: PatternFill = FILL_WHITE,
        font_overrides: Optional[Dict[int, Font]] = None,
    ):
        self.fill = fill
        self.font_overrides = font_overrides or {}


# ─── Generic sheet writer ────────────────────────────────────────────────────

def write_data_sheet(
    ws: Worksheet,
    columns: Sequence[ColSpec],
    rows: Sequence[Sequence[Any]],
    *,
    row_styles: Optional[Sequence[RowStyle]] = None,
    include_hidden_id: bool = False,
    ids: Optional[Sequence[str]] = None,
    freeze: bool = True,
) -> None:
    """
    Populate a worksheet with a header row and data rows.

    Args:
        ws: Target worksheet (already created).
        columns: Column definitions (header text + width).
        rows: 2-D data, one inner sequence per row.  Length of each inner
              sequence must match ``len(columns)``.
        row_styles: Optional per-row styling.  If shorter than *rows*,
                    remaining rows use defaults.
        include_hidden_id: If True, column A is a hidden ``__field_id__``
                           column and visible columns start at B.
        ids: When *include_hidden_id* is True, one id string per row.
        freeze: Freeze the header row.
    """
    col_offset = 1  # visible columns start at column 1 by default

    if include_hidden_id:
        ws.column_dimensions["A"].hidden = True
        ws.cell(row=1, column=1, value="__field_id__")
        col_offset = 2  # visible columns start at B

    # Header row
    for ci, col in enumerate(columns, start=col_offset):
        cell = ws.cell(row=1, column=ci, value=col.header)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

    # Data rows
    for ri, row_data in enumerate(rows, start=2):
        if include_hidden_id and ids:
            id_cell = ws.cell(row=ri, column=1, value=ids[ri - 2] if ri - 2 < len(ids) else "")
            id_cell.font = FONT_HIDDEN

        style = (
            row_styles[ri - 2]
            if row_styles and ri - 2 < len(row_styles)
            else RowStyle()
        )

        for ci, value in enumerate(row_data, start=col_offset):
            cell = ws.cell(row=ri, column=ci, value=value)
            cell.fill = style.fill
            cell.font = style.font_overrides.get(ci, FONT_DEFAULT)
            col_spec = columns[ci - col_offset]
            cell.alignment = ALIGN_CENTER if col_spec.align == "center" else ALIGN_LEFT
            cell.border = THIN_BORDER

    # Column widths
    for ci, col in enumerate(columns, start=col_offset):
        letter = ws.cell(row=1, column=ci).column_letter
        ws.column_dimensions[letter].width = col.width

    if freeze:
        freeze_col = "B" if include_hidden_id else "A"
        ws.freeze_panes = f"{freeze_col}2"


def write_kv_sheet(
    ws: Worksheet,
    pairs: Sequence[Tuple[str, Any]],
    *,
    label_width: int = 25,
    value_width: int = 60,
    row_fills: Optional[Dict[int, PatternFill]] = None,
    row_fonts: Optional[Dict[int, Font]] = None,
) -> None:
    """
    Write a simple two-column key/value sheet.

    Args:
        ws: Target worksheet.
        pairs: (label, value) tuples.
        label_width / value_width: Column widths.
        row_fills: Optional per-row-index fill overrides.
        row_fonts: Optional per-row-index value font overrides.
    """
    ws.column_dimensions["A"].width = label_width
    ws.column_dimensions["B"].width = value_width

    for ri, (label, value) in enumerate(pairs, start=1):
        lbl_cell = ws.cell(row=ri, column=1, value=label)
        lbl_cell.font = FONT_BOLD
        lbl_cell.alignment = ALIGN_LEFT
        lbl_cell.border = THIN_BORDER

        if isinstance(value, (dict, list)):
            value = str(value)
        val_cell = ws.cell(row=ri, column=2, value=value)
        val_cell.font = row_fonts.get(ri, FONT_DEFAULT) if row_fonts else FONT_DEFAULT
        val_cell.fill = row_fills.get(ri, FILL_WHITE) if row_fills else FILL_WHITE
        val_cell.alignment = ALIGN_LEFT
        val_cell.border = THIN_BORDER


def write_meta_row(
    ws: Worksheet,
    data_row_count: int,
    case_id: str,
    version: int,
    extra: Optional[Dict[str, str]] = None,
) -> None:
    """Append a hidden metadata row below the data for import validation."""
    meta_row = data_row_count + 3
    ws.cell(row=meta_row, column=1, value="__meta__")
    ws.cell(row=meta_row, column=2, value=f"case_id={case_id}")
    ws.cell(row=meta_row, column=3, value=f"version={version}")
    col = 4
    if extra:
        for k, v in extra.items():
            ws.cell(row=meta_row, column=col, value=f"{k}={v}")
            col += 1
    ws.row_dimensions[meta_row].hidden = True


def workbook_to_bytes(wb: Workbook) -> bytes:
    """Serialize an openpyxl Workbook to bytes."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
