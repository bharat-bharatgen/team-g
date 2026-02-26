"""
MER Excel Import Service

Reads an edited .xlsx, compares against the previous version's fields,
and produces a new MERResultModel snapshot.

Rules:
- If a field's answer/details changed → source="user", confidence=1.0
- If unchanged → keep original source & confidence
- Fields are matched by the hidden __field_id__ column
"""

import io
from datetime import datetime
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

from app.models.mer_result import MERField, MERResultModel, FieldSource


def _parse_excel(file_bytes: bytes) -> Tuple[List[dict], dict]:
    """
    Parse an uploaded Excel file and extract field rows + metadata.

    Returns:
        (rows, metadata)
        rows: List of dicts with keys: id, page, section, key, answer, details, confidence, source
        metadata: dict with case_id, version, fields_count
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    rows = []
    metadata = {}

    for row in ws.iter_rows(min_row=2, values_only=False):
        # Column A = hidden field id
        field_id = row[0].value

        if field_id == "__meta__":
            # Parse metadata row
            for cell in row[1:]:
                if cell.value and "=" in str(cell.value):
                    k, v = str(cell.value).split("=", 1)
                    metadata[k.strip()] = v.strip()
            continue

        if field_id is None or str(field_id).startswith("__"):
            continue

        # Columns B-H map to: Page, Section, Field, Answer, Details, Confidence, Source
        rows.append({
            "id": str(field_id),
            "page": int(row[1].value) if row[1].value else 0,
            "section": str(row[2].value or ""),
            "key": str(row[3].value or ""),
            "answer": str(row[4].value) if row[4].value else None,
            "details": str(row[5].value) if row[5].value else None,
            "confidence": float(row[6].value) if (row[6].value is not None and str(row[6].value).strip() != "") else None,
            "source": str(row[7].value or "llm"),
        })

    wb.close()
    return rows, metadata


def _normalize(val: Optional[str]) -> str:
    """Normalize a field value for comparison (strip whitespace, lowercase)."""
    if val is None:
        return ""
    return str(val).strip()


def _build_updated_fields(
    excel_rows: List[dict],
    prev_fields: List[MERField],
) -> List[MERField]:
    """
    Compare Excel rows against previous fields and produce updated field list.

    Changed fields → source="user", confidence=1.0
    Unchanged fields → keep original source & confidence
    """
    # Index previous fields by id for fast lookup
    prev_by_id: Dict[str, MERField] = {f.id: f for f in prev_fields}

    updated_fields: List[MERField] = []

    for row in excel_rows:
        field_id = row["id"]
        prev = prev_by_id.get(field_id)

        if prev is None:
            # Field id not found in previous version — treat as user-added
            updated_fields.append(MERField(
                id=field_id,
                page=row["page"],
                section=row["section"],
                key=row["key"],
                answer=row["answer"],
                details=row["details"],
                confidence=1.0,
                source=FieldSource.USER,
            ))
            continue

        # Check if user modified any content field
        answer_changed = _normalize(row["answer"]) != _normalize(prev.answer)
        details_changed = _normalize(row["details"]) != _normalize(prev.details)

        if answer_changed or details_changed:
            # User edited this field
            updated_fields.append(MERField(
                id=field_id,
                page=row["page"],
                section=row["section"],
                key=row["key"],
                answer=row["answer"],
                details=row["details"],
                confidence=1.0,
                source=FieldSource.USER,
            ))
        else:
            # Unchanged — keep original
            updated_fields.append(prev.model_copy())

    return updated_fields


def import_excel(
    file_bytes: bytes,
    prev_result: MERResultModel,
) -> MERResultModel:
    """
    Process an uploaded Excel file and create a new MERResultModel snapshot.

    Args:
        file_bytes: Raw bytes of the uploaded .xlsx file
        prev_result: The previous MERResultModel version this Excel was exported from

    Returns:
        A new MERResultModel with incremented version.
        The caller is responsible for storing it in MongoDB.
    """
    excel_rows, metadata = _parse_excel(file_bytes)

    # Validate metadata
    expected_case = prev_result.case_id
    if metadata.get("case_id") and metadata["case_id"] != expected_case:
        raise ValueError(
            f"Excel case_id mismatch: got {metadata.get('case_id')}, expected {expected_case}"
        )

    # Build updated fields
    updated_fields = _build_updated_fields(excel_rows, prev_result.fields)

    # Create new snapshot
    new_result = MERResultModel(
        case_id=prev_result.case_id,
        version=prev_result.version + 1,
        source="excel_import",
        classification=prev_result.classification,
        pages=prev_result.pages,
        fields=updated_fields,
        created_at=datetime.utcnow(),
    )

    return new_result
