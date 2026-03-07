"""
Risk Analysis Excel Export Service.

Generates an .xlsx from a risk analysis result document.
All LLM findings are consolidated into a single "Risk Analysis" sheet
with section headers, matching the frontend layout.

Pre-computed flags/contradictions get their own sheets if present.
"""

from typing import Any, Dict, List

from openpyxl import Workbook

from app.services.excel_utils import (
    FILL_GREEN,
    FILL_HEADER,
    FILL_LIGHT_BLUE,
    FILL_ORANGE,
    FILL_RED,
    FILL_WHITE,
    FILL_YELLOW,
    FONT_BOLD,
    FONT_DEFAULT,
    FONT_HEADER,
    ALIGN_CENTER,
    ALIGN_LEFT,
    THIN_BORDER,
    ColSpec,
    RowStyle,
    workbook_to_bytes,
    write_data_sheet,
    write_meta_row,
)


_SEVERITY_FILLS = {
    "critical": FILL_RED,
    "moderate": FILL_YELLOW,
    "mild": FILL_LIGHT_BLUE,
}

_RISK_LEVEL_FILLS = {
    "High": FILL_RED,
    "Intermediate": FILL_YELLOW,
    "Low": FILL_GREEN,
}

# Max visible columns used across all inline sections (for merged section headers)
_MAX_COLS = 3


def _refs_str(refs: Any) -> str:
    if isinstance(refs, list):
        return ", ".join(str(r) for r in refs)
    return str(refs) if refs else ""


def _is_v2(llm_response: dict) -> bool:
    return "integrity_concerns" in llm_response


# ─── Single-sheet builder ────────────────────────────────────────────────────

def _write_section_header(ws, row: int, title: str) -> int:
    """Write a merged section header row. Returns the next available row."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_MAX_COLS)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = FONT_HEADER
    cell.fill = FILL_HEADER
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER
    for col in range(2, _MAX_COLS + 1):
        c = ws.cell(row=row, column=col)
        c.fill = FILL_HEADER
        c.border = THIN_BORDER
    return row + 1


def _write_kv_row(ws, row: int, label: str, value: Any, fill=FILL_WHITE, value_font=FONT_DEFAULT) -> int:
    """Write a label-value pair spanning columns 1-3. Returns next row."""
    lbl = ws.cell(row=row, column=1, value=label)
    lbl.font = FONT_BOLD
    lbl.alignment = ALIGN_LEFT
    lbl.border = THIN_BORDER

    if isinstance(value, (dict, list)):
        value = str(value)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=_MAX_COLS)
    val = ws.cell(row=row, column=2, value=value)
    val.font = value_font
    val.fill = fill
    val.alignment = ALIGN_LEFT
    val.border = THIN_BORDER
    for col in range(3, _MAX_COLS + 1):
        c = ws.cell(row=row, column=col)
        c.border = THIN_BORDER
    return row + 1


def _write_table_header(ws, row: int, headers: List[str]) -> int:
    """Write a table column header row. Returns next row."""
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = FONT_BOLD
        cell.fill = FILL_WHITE
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
    return row + 1


def _write_table_row(ws, row: int, values: List[Any], fill=FILL_WHITE, font=FONT_DEFAULT) -> int:
    """Write a single data row. Returns next row."""
    for ci, v in enumerate(values, start=1):
        cell = ws.cell(row=row, column=ci, value=v)
        cell.font = font
        cell.fill = fill
        cell.alignment = ALIGN_LEFT if ci == 1 else ALIGN_CENTER
        cell.border = THIN_BORDER
    return row + 1


def _build_main_sheet(ws, doc: dict) -> None:
    """Build the consolidated Risk Analysis sheet."""
    llm = doc.get("llm_response", {})
    risk_level = llm.get("risk_level", "Unknown")
    v2 = _is_v2(llm)
    row = 1

    # ── Section: Summary ─────────────────────────────────────────
    row = _write_section_header(ws, row, "Summary")

    risk_fill = _RISK_LEVEL_FILLS.get(risk_level, FILL_WHITE)
    row = _write_kv_row(ws, row, "Risk Level", risk_level, fill=risk_fill, value_font=FONT_BOLD)

    if v2:
        row = _write_kv_row(ws, row, "Risk Score", f"{llm.get('risk_score', 'N/A')} / 10")
        row = _write_kv_row(ws, row, "Applicant", llm.get("applicant", "N/A"))

    summary = llm.get("summary", "")
    if isinstance(summary, dict):
        label_map = {"mer": "MER Findings", "pathology": "Pathology Findings", "conclusion": "Conclusion"}
        for key, val in summary.items():
            label = label_map.get(key, key.replace("_", " ").title())
            row = _write_kv_row(ws, row, label, str(val))
    else:
        row = _write_kv_row(ws, row, "Summary", str(summary))

    row = _write_kv_row(ws, row, "Generated At", str(doc.get("created_at", "")))
    row += 1  # blank separator

    # ── Section: Integrity Concerns / Red Flags ──────────────────
    if v2:
        row = _write_section_header(ws, row, "Integrity Concerns")
        integrity = llm.get("integrity_concerns") or []
        if integrity:
            row = _write_table_header(ws, row, ["Flag", "MER Reference", "Pathology Reference"])
            for item in integrity:
                row = _write_table_row(
                    ws, row,
                    [item.get("flag", ""), item.get("mer_ref", ""), item.get("path_ref", "")],
                    fill=FILL_RED,
                )
        else:
            row = _write_table_row(ws, row, ["None", "", ""])
    else:
        row = _write_section_header(ws, row, "Red Flags")
        red_flags = llm.get("red_flags") or []
        if red_flags:
            row = _write_table_header(ws, row, ["Red Flag", "References", ""])
            for item in red_flags:
                text = item if isinstance(item, str) else item.get("text", "")
                refs = "" if isinstance(item, str) else _refs_str(item.get("refs", []))
                row = _write_table_row(ws, row, [text, refs, ""], fill=FILL_RED)
        else:
            row = _write_table_row(ws, row, ["None", "", ""])

    row += 1  # blank separator

    # ── Section: Clinical Discoveries / Contradictions ────────────
    if v2:
        row = _write_section_header(ws, row, "Clinical Discoveries")
        clinical = llm.get("clinical_discoveries") or []
        if clinical:
            row = _write_table_header(ws, row, ["Finding", "Severity", "References"])
            for item in clinical:
                severity = item.get("severity", "")
                fill = _SEVERITY_FILLS.get(severity, FILL_WHITE)
                row = _write_table_row(
                    ws, row,
                    [item.get("finding", ""), severity, _refs_str(item.get("refs", []))],
                    fill=fill,
                )
        else:
            row = _write_table_row(ws, row, ["None", "", ""])
    else:
        row = _write_section_header(ws, row, "Contradictions")
        contras = llm.get("contradictions") or []
        if contras:
            row = _write_table_header(ws, row, ["Contradiction", "References", ""])
            for item in contras:
                text = item if isinstance(item, str) else item.get("text", "")
                refs = "" if isinstance(item, str) else _refs_str(item.get("refs", []))
                row = _write_table_row(ws, row, [text, refs, ""], fill=FILL_ORANGE)
        else:
            row = _write_table_row(ws, row, ["None", "", ""])

    # ── Column widths ────────────────────────────────────────────
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 25


# ─── Pre-computed sheet builders (separate sheets) ───────────────────────────

def _build_precomputed_flags_sheet(ws, flags: List[Dict]) -> None:
    columns = [
        ColSpec("Parameter", width=20),
        ColSpec("Value", width=18),
        ColSpec("Source", width=12, align="center"),
        ColSpec("Severity", width=12, align="center"),
        ColSpec("Message", width=50),
    ]
    rows = [
        [f.get("parameter", ""), f.get("value", ""), f.get("source", ""), f.get("severity", ""), f.get("message", "")]
        for f in flags
    ]
    styles = [RowStyle(fill=FILL_YELLOW) for _ in flags]
    write_data_sheet(ws, columns, rows, row_styles=styles)


def _build_precomputed_contradictions_sheet(ws, items: List[Dict]) -> None:
    columns = [
        ColSpec("Field", width=20),
        ColSpec("Type", width=20),
        ColSpec("MER Value", width=25),
        ColSpec("Pathology Value", width=25),
        ColSpec("Severity", width=12, align="center"),
    ]
    rows = [
        [c.get("field", ""), c.get("type", ""), c.get("mer_value", ""), c.get("pathology_value", ""), c.get("severity", "")]
        for c in items
    ]
    styles = [RowStyle(fill=FILL_ORANGE) for _ in items]
    write_data_sheet(ws, columns, rows, row_styles=styles)


# ─── Public API ──────────────────────────────────────────────────────────────

def generate_excel(doc: dict, case_id: str, version: int) -> bytes:
    """
    Generate an Excel workbook from a risk analysis result document.

    Main sheet: "Risk Analysis" — consolidated summary + findings.
    Additional sheets for pre-computed flags/contradictions if present.
    """
    wb = Workbook()

    ws_main = wb.active
    ws_main.title = "Risk Analysis"
    _build_main_sheet(ws_main, doc)

    precomp_flags = doc.get("critical_flags") or []
    if precomp_flags:
        ws = wb.create_sheet("Critical Values")
        _build_precomputed_flags_sheet(ws, precomp_flags)

    precomp_contras = doc.get("contradictions") or []
    if precomp_contras:
        ws = wb.create_sheet("Data Contradictions")
        _build_precomputed_contradictions_sheet(ws, precomp_contras)

    write_meta_row(ws_main, 30, case_id, version)

    return workbook_to_bytes(wb)
