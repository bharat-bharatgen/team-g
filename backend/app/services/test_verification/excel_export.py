"""
Test Verification Excel Export Service.

Generates an .xlsx from a TestVerificationResultModel document.
All sections consolidated into a single "Test Verification" sheet
with section headers, matching the frontend layout.

Coloring:
    | Condition         | Color     |
    |-------------------|-----------|
    | found == True     | White     |
    | found == False    | Red       |
    | status: complete  | Green     |
    | status: missing   | Red       |
"""

from typing import Any, Dict, List

from openpyxl import Workbook

from app.services.excel_utils import (
    FILL_GREEN,
    FILL_HEADER,
    FILL_RED,
    FILL_WHITE,
    FONT_BOLD,
    FONT_DEFAULT,
    FONT_HEADER,
    ALIGN_CENTER,
    ALIGN_LEFT,
    THIN_BORDER,
    workbook_to_bytes,
    write_meta_row,
)

_MAX_COLS = 4


def _write_section_header(ws, row: int, title: str) -> int:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_MAX_COLS)
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = FONT_HEADER
    cell.fill = FILL_HEADER
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER
    for col in range(2, _MAX_COLS + 1):
        c = ws.cell(row=row, column=col)
        c.fill = FILL_HEADER
        c.border = THIN_BORDER
    return row + 1


def _write_kv_row(ws, row: int, label: str, value: Any, fill=FILL_WHITE, value_font=FONT_DEFAULT) -> int:
    lbl = ws.cell(row=row, column=1, value=label)
    lbl.font = FONT_BOLD
    lbl.alignment = ALIGN_LEFT
    lbl.border = THIN_BORDER

    if isinstance(value, (dict, list)):
        value = str(value)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=_MAX_COLS)
    val = ws.cell(row=row, column=2, value=value)
    val.font = value_font
    val.fill = fill
    val.alignment = ALIGN_LEFT
    val.border = THIN_BORDER
    for col in range(3, _MAX_COLS + 1):
        c = ws.cell(row=row, column=col)
        c.border = THIN_BORDER
    return row + 1


def _write_table_header(ws, row: int, headers: List[str]) -> int:
    for ci, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = FONT_BOLD
        cell.fill = FILL_WHITE
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
    return row + 1


def _write_table_row(ws, row: int, values: List[Any], fill=FILL_WHITE, font=FONT_DEFAULT) -> int:
    for ci, v in enumerate(values, start=1):
        cell = ws.cell(row=row, column=ci, value=v)
        cell.font = font
        cell.fill = fill
        cell.alignment = ALIGN_LEFT if ci <= 2 else ALIGN_CENTER
        cell.border = THIN_BORDER
    return row + 1


def generate_excel(doc: dict, case_id: str, version: int) -> bytes:
    """
    Generate an Excel workbook from a test verification result document.
    Single sheet with Summary, Required Tests, and Missing Tests sections.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Verification"
    row = 1

    status = doc.get("status", "unknown")
    total_req = doc.get("total_required", 0)
    total_found = doc.get("total_found", 0)
    total_missing = doc.get("total_missing", 0)
    pct = f"{round(total_found / total_req * 100)}%" if total_req > 0 else "N/A"
    status_fill = FILL_GREEN if status == "complete" else FILL_RED if status == "missing_tests" else FILL_WHITE

    # ── Section: Summary ─────────────────────────────────────────
    row = _write_section_header(ws, row, "Summary")
    row = _write_kv_row(ws, row, "Status", status.replace("_", " ").title(), fill=status_fill, value_font=FONT_BOLD)
    row = _write_kv_row(ws, row, "Total Required", total_req)
    row = _write_kv_row(ws, row, "Total Found", total_found)
    row = _write_kv_row(ws, row, "Total Missing", total_missing)
    row = _write_kv_row(ws, row, "Completion", pct)

    proposal = doc.get("proposal_number")
    if proposal:
        row = _write_kv_row(ws, row, "Proposal Number", proposal)

    name = doc.get("life_assured_name")
    if name:
        row = _write_kv_row(ws, row, "Life Assured", name)

    ins_remark = doc.get("ins_test_remark")
    if ins_remark:
        row = _write_kv_row(ws, row, "Insurance Test Remark", ins_remark)

    confidence = doc.get("extraction_confidence", 0)
    if confidence > 0:
        row = _write_kv_row(ws, row, "Extraction Confidence", f"{round(confidence * 100)}%")

    row = _write_kv_row(ws, row, "Generated At", str(doc.get("created_at", "")))
    row += 1

    # ── Section: Required Tests ──────────────────────────────────
    tests: List[Dict[str, Any]] = doc.get("required_tests") or []
    row = _write_section_header(ws, row, "Required Tests")
    if tests:
        row = _write_table_header(ws, row, ["Category", "Test Name", "Status", "Pathology Value"])
        for t in tests:
            found = t.get("found", False)
            fill = FILL_WHITE if found else FILL_RED
            row = _write_table_row(
                ws, row,
                [
                    t.get("category", ""),
                    t.get("test_name", ""),
                    "Found" if found else "Missing",
                    t.get("pathology_value") or ("Present" if found else "Not Found"),
                ],
                fill=fill,
            )
    else:
        row = _write_table_row(ws, row, ["No required tests found", "", "", ""])
    row += 1

    # ── Section: Missing Tests ───────────────────────────────────
    missing: List[str] = doc.get("missing_tests") or []
    row = _write_section_header(ws, row, "Missing Tests")
    if missing:
        for m in missing:
            row = _write_table_row(ws, row, [m, "", "", ""], fill=FILL_RED)
    else:
        row = _write_table_row(ws, row, ["None", "", "", ""])

    # ── Column widths ────────────────────────────────────────────
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 25

    # ── Metadata ─────────────────────────────────────────────────
    write_meta_row(ws, row + 2, case_id, version, extra={"tests_count": str(len(tests))})

    return workbook_to_bytes(wb)
