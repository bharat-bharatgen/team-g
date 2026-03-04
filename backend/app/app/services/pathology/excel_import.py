"""
Pathology Excel Import Service

Reads an edited .xlsx, compares against the previous version's fields,
and produces a new PathologyResultModel snapshot.

Rules:
- If a field's value/unit/reference_range/method changed → source="user"
- If unchanged → keep original source
- Fields are matched by the hidden __field_id__ column
- range_status is recomputed after import
"""

import io
from datetime import datetime
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook

from app.models.pathology_result import PathologyField, PathologyResultModel, FieldSource
from app.services.pathology.range_utils import compute_range_status


# Column indices (0-based, after hidden column A at index 0)
# Columns: A=id, B=Parameter, C=Original Name, D=Value, E=Unit, F=Report Range, G=Standard Range, H=Status, I=Method, J=Is Standard, K=Source
COL_ID = 0
COL_PARAMETER = 1
COL_ORIGINAL_NAME = 2
COL_VALUE = 3
COL_UNIT = 4
COL_REPORT_RANGE = 5
COL_STANDARD_RANGE = 6
COL_STATUS = 7
COL_METHOD = 8
COL_IS_STANDARD = 9
COL_SOURCE = 10


def _parse_excel(file_bytes: bytes) -> Tuple[List[dict], dict]:
    """
    Parse an uploaded Excel file and extract field rows + metadata.

    Returns:
        (rows, metadata)
        rows: List of dicts with field data
        metadata: dict with case_id, version, fields_count
    """
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    rows = []
    metadata = {}

    for row in ws.iter_rows(min_row=2, values_only=False):
        # Column A = hidden field id
        field_id = row[COL_ID].value

        if field_id == "__meta__":
            # Parse metadata row
            for cell in row[1:]:
                if cell.value and "=" in str(cell.value):
                    k, v = str(cell.value).split("=", 1)
                    metadata[k.strip()] = v.strip()
            continue

        if field_id is None or str(field_id).startswith("__"):
            continue

        # Parse row data
        is_std_val = str(row[COL_IS_STANDARD].value or "Yes").strip().lower()
        rows.append({
            "id": str(field_id),
            "key": str(row[COL_PARAMETER].value or ""),
            "reference_name": str(row[COL_ORIGINAL_NAME].value) if row[COL_ORIGINAL_NAME].value else None,
            "value": str(row[COL_VALUE].value) if row[COL_VALUE].value else None,
            "unit": str(row[COL_UNIT].value) if row[COL_UNIT].value else None,
            "reference_range": str(row[COL_REPORT_RANGE].value) if row[COL_REPORT_RANGE].value else None,
            "config_range": str(row[COL_STANDARD_RANGE].value) if row[COL_STANDARD_RANGE].value else None,
            "method": str(row[COL_METHOD].value) if row[COL_METHOD].value else None,
            "is_standard": is_std_val in ("yes", "true", "1"),
            "source": str(row[COL_SOURCE].value or "llm"),
        })

    wb.close()
    return rows, metadata


def _normalize(val: Optional[str]) -> str:
    """Normalize a field value for comparison (strip whitespace)."""
    if val is None:
        return ""
    return str(val).strip()


def _build_updated_fields(
    excel_rows: List[dict],
    prev_fields: List[PathologyField],
) -> List[PathologyField]:
    """
    Compare Excel rows against previous fields and produce updated field list.

    Changed fields → source="user"
    Unchanged fields → keep original source
    Recomputes range_status for all fields.
    """
    # Index previous fields by id for fast lookup
    prev_by_id: Dict[str, PathologyField] = {f.id: f for f in prev_fields}

    updated_fields: List[PathologyField] = []

    for row in excel_rows:
        field_id = row["id"]
        prev = prev_by_id.get(field_id)

        value = row["value"]
        reference_range = row["reference_range"]
        config_range = row.get("config_range")

        # Recompute range_status
        range_status = compute_range_status(value, reference_range, config_range)

        if prev is None:
            # Field id not found in previous version — treat as user-added
            updated_fields.append(PathologyField(
                id=field_id,
                key=row["key"],
                value=value,
                unit=row["unit"],
                reference_range=reference_range,
                config_range=config_range,
                range_status=range_status,
                method=row["method"],
                reference_name=row.get("reference_name"),
                is_standard=row["is_standard"],
                source=FieldSource.USER,
            ))
            continue

        # Check if user modified any content field
        value_changed = _normalize(value) != _normalize(prev.value)
        unit_changed = _normalize(row["unit"]) != _normalize(prev.unit)
        range_changed = _normalize(reference_range) != _normalize(prev.reference_range)
        method_changed = _normalize(row["method"]) != _normalize(prev.method)

        if value_changed or unit_changed or range_changed or method_changed:
            # User edited this field
            updated_fields.append(PathologyField(
                id=field_id,
                key=row["key"],
                value=value,
                unit=row["unit"],
                reference_range=reference_range,
                config_range=config_range or prev.config_range,
                range_status=range_status,
                method=row["method"],
                reference_name=row.get("reference_name") or prev.reference_name,
                sample_type=prev.sample_type,
                section_path=prev.section_path,
                is_standard=row["is_standard"],
                source=FieldSource.USER,
            ))
        else:
            # Unchanged — keep original but update range_status if needed
            updated = prev.model_copy()
            updated.range_status = range_status
            updated_fields.append(updated)

    return updated_fields


def import_excel(
    file_bytes: bytes,
    prev_result: PathologyResultModel,
) -> PathologyResultModel:
    """
    Process an uploaded Excel file and create a new PathologyResultModel snapshot.

    Args:
        file_bytes: Raw bytes of the uploaded .xlsx file
        prev_result: The previous PathologyResultModel version this Excel was exported from

    Returns:
        A new PathologyResultModel with incremented version.
        The caller is responsible for storing it in MongoDB.
    """
    excel_rows, metadata = _parse_excel(file_bytes)

    # Validate metadata
    expected_case = prev_result.case_id
    if metadata.get("case_id") and metadata["case_id"] != expected_case:
        raise ValueError(
            f"Excel case_id mismatch: got {metadata.get('case_id')}, expected {expected_case}"
        )

    # Build updated fields
    updated_fields = _build_updated_fields(excel_rows, prev_result.fields)

    # Create new snapshot
    new_result = PathologyResultModel(
        case_id=prev_result.case_id,
        version=prev_result.version + 1,
        source="excel_import",
        pages=prev_result.pages,
        patient_info=prev_result.patient_info,
        lab_info=prev_result.lab_info,
        report_info=prev_result.report_info,
        standardized=prev_result.standardized,
        fields=updated_fields,
        created_at=datetime.utcnow(),
    )

    return new_result
